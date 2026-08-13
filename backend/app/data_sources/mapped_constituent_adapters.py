"""需由官方目錄對照內部基金 ID 的成分股來源。"""

import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Callable
from urllib.parse import urljoin
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook

from backend.app.data_sources.direct_constituent_adapters import (
    USER_AGENT,
    _TableParser,
    _normalize_code,
    _parse_date,
    _positions,
    _snapshot,
)
from backend.app.data_sources.fuh_hwa_actual_dividend_adapter import (
    parse_fuh_hwa_internal_id,
)
from backend.app.data_sources.openapi import create_ssl_context
from backend.app.data_sources.uob_actual_dividend_adapter import (
    parse_uob_fund_detail_url,
)
from backend.app.models.etf_constituent import ETFConstituentSnapshotCreate


MEGA_CATALOG_URL = "https://www.megafunds.com.tw/MEGA/etf/"
MEGA_BASE_URL = "https://www.megafunds.com.tw/MEGA/etf/"
FUH_HWA_CATALOG_URL = "https://www.fhtrust.com.tw/ETF/etf_detail/ETF01"
FUH_HWA_BASE_URL = "https://www.fhtrust.com.tw"
CAPITAL_CATALOG_URL = "https://www.capitalfund.com.tw/etf/product"
CAPITAL_BASE_URL = "https://www.capitalfund.com.tw"
UOB_EVENT_URL = "https://www.uobam.com.tw/Events/{etf_code}ETF/"
UOB_BASE_URL = "https://www.uobam.com.tw"
ESUN_API_BASE_URL = "https://www.esunam.com/ETFAPI"
ESUN_SOURCE_URL = "https://www.esunam.com/ETF/etf-pcf"
FRANKLIN_API_BASE_URL = "https://www.ftft.com.tw/official/api"
FRANKLIN_SOURCE_URL = "https://www.ftft.com.tw/etf/product/details/"
MAX_RESPONSE_BYTES = 5_000_000


def _get(url: str, *, timeout_seconds: float = 30) -> httpx.Response:
    response = httpx.get(
        url, timeout=timeout_seconds, follow_redirects=True,
        verify=create_ssl_context(allow_legacy_x509=True),
        headers={"User-Agent": USER_AGENT},
    )
    response.raise_for_status()
    if len(response.content) > MAX_RESPONSE_BYTES:
        raise ValueError("官方成分股回應超過容量上限")
    return response


def _post_json(url: str, payload: dict[str, Any], *, timeout_seconds: float = 30) -> dict:
    response = httpx.post(
        url, json=payload, timeout=timeout_seconds, follow_redirects=True,
        verify=create_ssl_context(allow_legacy_x509=True),
        headers={"Accept": "application/json", "User-Agent": USER_AGENT},
    )
    response.raise_for_status()
    if len(response.content) > MAX_RESPONSE_BYTES:
        raise ValueError("官方成分股回應超過容量上限")
    value = response.json()
    if not isinstance(value, dict):
        raise ValueError("官方成分股 API 未回傳物件")
    return value


def parse_mega_product_url(*, etf_code: str, catalog_html: str) -> str:
    code = _normalize_code(etf_code)
    pattern = rf'<div[^>]+class="detail-item"[^>]*>\s*{re.escape(code)}\s*</div>.*?href="([^"]*etf_product\.aspx\?id=\d+)"'
    match = re.search(pattern, catalog_html, re.I | re.S)
    if match is None:
        raise ValueError(f"兆豐官方 ETF 總覽找不到證券代號：{code}")
    return urljoin(MEGA_BASE_URL, match.group(1))


def parse_mega_constituent_html(
    content: str, *, etf_code: str, source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    if code not in content.upper():
        raise ValueError(f"兆豐官方持股回應與要求的 {code} 不符")
    date_match = re.search(r"資料來源：兆豐投信，(20\d{2}/\d{1,2}/\d{1,2})", content)
    start = content.find('id="fund_content_list_1"')
    end = content.find('id="fund-content-2"', start)
    if start < 0 or end < 0:
        raise ValueError("兆豐官方持股找不到股票權重表")
    rows = re.findall(
        r'class="fund-info content-list-1"[^>]*>(.*?)(?=<div[^>]+class="fund-info content-list-1"|$)',
        content[start:end], re.I | re.S,
    )
    table = [["股票代號", "股票名稱", "股數", "持股權重"]]
    for row in rows:
        cells = [re.sub(r"<[^>]+>", "", value).strip() for value in re.findall(
            r'<div[^>]+class="fund-content[^"]*"[^>]*>(.*?)</div>', row, re.I | re.S
        )]
        if len(cells) >= 4:
            table.append(cells[:4])
    return _snapshot(
        code, "兆豐", "mega_official_holdings", source_url,
        _parse_date(date_match.group(1) if date_match else "", "兆豐"),
        _positions(table, code_index=0, name_index=1, weight_index=3, issuer="兆豐"),
        fetched_at,
    )


def parse_fuh_hwa_assets_link(*, detail_html: str, internal_id: str) -> str:
    match = re.search(
        rf'href="(/api/assetsExcel/{re.escape(internal_id)}/(20\d{{6}}))"',
        detail_html, re.I,
    )
    if match is None:
        raise ValueError("復華官方基金頁找不到資產 Excel")
    return urljoin(FUH_HWA_BASE_URL, match.group(1))


def parse_fuh_hwa_constituent_excel(
    content: bytes, *, etf_code: str, source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    try:
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        rows = [["" if value is None else str(value).strip() for value in row]
                for row in sheet.iter_rows(values_only=True)]
    except Exception as error:
        raise ValueError("復華官方資產 Excel 無法解析") from error
    if not rows or code not in rows[0][0].upper():
        raise ValueError(f"復華官方持股回應與要求的 {code} 不符")
    date_row = next((row[0] for row in rows if row and row[0].startswith("日期:")), "")
    header_index = next(
        (index for index, row in enumerate(rows) if row[:2] == ["證券代號", "證券名稱"]),
        None,
    )
    if header_index is None:
        raise ValueError("復華官方資產 Excel 找不到股票權重表")
    table = [["證券代號", "證券名稱", "股數", "金額", "權重(%)"]]
    for row in rows[header_index + 1:]:
        if len(row) < 5 or not row[0] or row[0] in {"股票合計", "期貨合計"}:
            continue
        if not re.fullmatch(r"[0-9A-Z. -]{2,30}", row[0].upper()):
            break
        table.append(row[:5])
    return _snapshot(
        code, "復華", "fuh_hwa_official_assets_excel", source_url,
        _parse_date(date_row, "復華"),
        _positions(table, code_index=0, name_index=1, weight_index=4, issuer="復華"),
        fetched_at,
    )


def parse_capital_product_url(*, etf_code: str, catalog_html: str) -> str:
    code = _normalize_code(etf_code)
    match = re.search(
        rf'href="/etf/product/detail/(\d+)/basic"[^>]*>\s*{re.escape(code)}\s*</a>',
        catalog_html, re.I,
    )
    if match is None:
        raise ValueError(f"群益官方 ETF 總覽找不到證券代號：{code}")
    return f"{CAPITAL_BASE_URL}/etf/product/detail/{match.group(1)}/buyback"


def parse_capital_constituent_html(
    content: str, *, etf_code: str, source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    if code not in content.upper():
        raise ValueError(f"群益官方持股回應與要求的 {code} 不符")
    stock_marker = content.find("股票代號")
    if stock_marker < 0:
        raise ValueError("群益官方持股找不到股票權重表")
    section = content[stock_marker:]
    date_matches = re.findall(r"\((20\d{2}/\d{1,2}/\d{1,2})\)", content[:stock_marker])
    rows = re.findall(
        r'class="tr show-for-medium"[^>]*>\s*<div[^>]+class="th"[^>]*>\s*([^<]+)</div>\s*<div[^>]+class="th"[^>]*>\s*([^<]+)</div>\s*<div[^>]+class="td"[^>]*>\s*([^<]+)</div>',
        section, re.I | re.S,
    )
    table = [["股票代號", "股票名稱", "持股權重(%)"], *[list(row) for row in rows]]
    return _snapshot(
        code, "群益", "capital_official_buyback", source_url,
        _parse_date(date_matches[-1] if date_matches else "", "群益"),
        _positions(table, code_index=0, name_index=1, weight_index=2, issuer="群益"),
        fetched_at,
    )


def parse_uob_constituent_html(
    content: str, *, etf_code: str, source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    if code not in content.upper():
        raise ValueError(f"大華銀官方持股回應與要求的 {code} 不符")
    date_match = re.search(r"資料日期\s*:\s*(20\d{2}/\d{1,2}/\d{1,2})", content)
    parser = _TableParser()
    parser.feed(content)
    table = next(
        (table for table in parser.tables if table and "股票代號" in table[0]
         and "佔基金淨資產之權重(%)" in table[0]),
        None,
    )
    if table is None:
        raise ValueError("大華銀官方持股找不到股票權重表")
    return _snapshot(
        code, "大華銀", "uob_official_pcf", source_url,
        _parse_date(date_match.group(1) if date_match else "", "大華銀"),
        _positions(table, code_index=0, name_index=1, weight_index=4, issuer="大華銀"),
        fetched_at,
    )


def parse_esun_mapping(*, etf_code: str, overview_payload: dict) -> str:
    code = _normalize_code(etf_code)
    if overview_payload.get("StatusCode") != 0 or not isinstance(overview_payload.get("Entries"), list):
        raise ValueError("玉山官方 ETF 總覽 API 回應失敗")
    fund = next(
        (row for row in overview_payload["Entries"]
         if str(row.get("StcokNo", "")).strip().upper() == code),
        None,
    )
    if not isinstance(fund, dict) or not str(fund.get("FundNo", "")).strip():
        raise ValueError(f"玉山官方 ETF 總覽找不到證券代號：{code}")
    return str(fund["FundNo"]).strip()


def parse_esun_constituent_payload(
    payload: dict, *, etf_code: str, fund_id: str,
    source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    entries = payload.get("Entries")
    if payload.get("StatusCode") != 0 or not isinstance(entries, dict):
        raise ValueError("玉山官方持股 API 回應失敗")
    if str(entries.get("FundID", "")) != fund_id:
        raise ValueError("玉山官方持股回傳基金 ID 不符")
    data = entries.get("Data")
    tables = data.get("Table") if isinstance(data, dict) else None
    stock = next(
        (row for row in tables or [] if isinstance(row, dict) and row.get("TableTitle") == "股票"),
        None,
    )
    asset = data.get("FundAsset") if isinstance(data, dict) else None
    if not isinstance(stock, dict) or not isinstance(stock.get("Rows"), list):
        raise ValueError("玉山官方持股缺少股票權重表")
    table = [["股票代號", "股票名稱", "股數", "權重(%)"], *stock["Rows"]]
    return _snapshot(
        code, "玉山", "esun_official_fund_assets", source_url,
        _parse_date(str(asset.get("NavDate", "")) if isinstance(asset, dict) else "", "玉山"),
        _positions(table, code_index=0, name_index=1, weight_index=3, issuer="玉山"),
        fetched_at,
    )


def parse_franklin_mapping(*, etf_code: str, catalog_payload: Any) -> str:
    """由富蘭克林官方 ETF 清單解析證券代號對應的基金 ID。"""

    code = _normalize_code(etf_code)
    if not isinstance(catalog_payload, list):
        raise ValueError("富蘭克林官方 ETF 清單回應格式錯誤")
    fund = next(
        (row for row in catalog_payload if isinstance(row, dict)
         and str(row.get("StockCode", "")).strip().upper() == code),
        None,
    )
    if not isinstance(fund, dict) or not str(fund.get("FundID", "")).strip():
        raise ValueError(f"富蘭克林官方 ETF 清單找不到證券代號：{code}")
    return str(fund["FundID"]).strip()


def parse_franklin_latest_query_date(date_payload: Any) -> str:
    """將官方 UTC 可選日期轉成台灣頁面實際查詢的 YYYYMMDD。"""

    if not isinstance(date_payload, list) or not date_payload:
        raise ValueError("富蘭克林官方持股缺少可用日期")
    try:
        values = [
            datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            for value in date_payload
        ]
    except ValueError as error:
        raise ValueError("富蘭克林官方持股包含無效日期") from error
    latest = max(values).astimezone(ZoneInfo("Asia/Taipei"))
    return latest.strftime("%Y%m%d")


def parse_franklin_constituent_payload(
    payload: Any, *, etf_code: str, fund_id: str,
    source_url: str, fetched_at: datetime,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    if not isinstance(payload, dict):
        raise ValueError("富蘭克林官方持股 API 回應格式錯誤")
    if str(payload.get("FundID", "")) != fund_id:
        raise ValueError("富蘭克林官方持股回傳基金 ID 不符")
    if str(payload.get("StockCode", "")).strip().upper() != code:
        raise ValueError(f"富蘭克林官方持股回應與要求的 {code} 不符")
    rows = payload.get("Secs")
    if not isinstance(rows, list) or not rows:
        raise ValueError("富蘭克林官方持股缺少股票權重")
    table = [["股票代號", "股票名稱", "權重(%)"]]
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("富蘭克林官方持股包含無效股票資料列")
        table.append([
            str(row.get("SecuritiesCode", "")).strip(),
            str(row.get("SecuritiesName", "")).strip(),
            str(row.get("WeightingPercentage", "")).strip(),
        ])
    asset_date = str(payload.get("AssetDate", ""))
    try:
        as_of_date = datetime.fromisoformat(
            asset_date.replace("Z", "+00:00")
        ).astimezone(ZoneInfo("Asia/Taipei")).date()
    except ValueError as error:
        raise ValueError("富蘭克林官方持股缺少有效資料日期") from error
    return _snapshot(
        code, "富蘭克林", "franklin_official_holdings_api", source_url,
        as_of_date,
        _positions(
            table, code_index=0, name_index=1, weight_index=2,
            issuer="富蘭克林",
        ),
        fetched_at,
    )
def fetch_mega_constituent_snapshot(etf_code: str, *, timeout_seconds: float = 30,
                                    fetched_at: datetime | None = None):
    code = _normalize_code(etf_code)
    catalog = _get(MEGA_CATALOG_URL, timeout_seconds=timeout_seconds)
    url = parse_mega_product_url(etf_code=code, catalog_html=catalog.text)
    return parse_mega_constituent_html(
        _get(url, timeout_seconds=timeout_seconds).text, etf_code=code,
        source_url=url, fetched_at=fetched_at or datetime.now(timezone.utc),
    )


def fetch_fuh_hwa_constituent_snapshot(etf_code: str, *, timeout_seconds: float = 30,
                                       fetched_at: datetime | None = None):
    code = _normalize_code(etf_code)
    catalog = _get(FUH_HWA_CATALOG_URL, timeout_seconds=timeout_seconds)
    internal_id = parse_fuh_hwa_internal_id(etf_code=code, html_text=catalog.text)
    detail_url = f"{FUH_HWA_BASE_URL}/ETF/etf_detail/{internal_id}"
    detail = _get(detail_url, timeout_seconds=timeout_seconds)
    assets_url = parse_fuh_hwa_assets_link(detail_html=detail.text, internal_id=internal_id)
    assets = _get(assets_url, timeout_seconds=timeout_seconds)
    return parse_fuh_hwa_constituent_excel(
        assets.content, etf_code=code, source_url=str(assets.url),
        fetched_at=fetched_at or datetime.now(timezone.utc),
    )


def fetch_capital_constituent_snapshot(etf_code: str, *, timeout_seconds: float = 30,
                                       fetched_at: datetime | None = None):
    code = _normalize_code(etf_code)
    catalog = _get(CAPITAL_CATALOG_URL, timeout_seconds=timeout_seconds)
    url = parse_capital_product_url(etf_code=code, catalog_html=catalog.text)
    return parse_capital_constituent_html(
        _get(url, timeout_seconds=timeout_seconds).text, etf_code=code,
        source_url=url, fetched_at=fetched_at or datetime.now(timezone.utc),
    )


def fetch_uob_constituent_snapshot(etf_code: str, *, timeout_seconds: float = 30,
                                   fetched_at: datetime | None = None):
    code = _normalize_code(etf_code)
    event = _get(UOB_EVENT_URL.format(etf_code=code), timeout_seconds=timeout_seconds)
    detail_url = parse_uob_fund_detail_url(etf_code=code, html_text=event.text)
    detail = _get(detail_url, timeout_seconds=timeout_seconds)
    match = re.search(r'href="/fund/etf/pcf\?fundID=(\d+)"', detail.text, re.I)
    if match is None:
        raise ValueError("大華銀官方基金頁找不到 PCF 基金 ID")
    pcf_url = f"{UOB_BASE_URL}/fund/etf/pcf?fundID={match.group(1)}"
    pcf = _get(pcf_url, timeout_seconds=timeout_seconds)
    return parse_uob_constituent_html(
        pcf.text, etf_code=code, source_url=str(pcf.url),
        fetched_at=fetched_at or datetime.now(timezone.utc),
    )


def fetch_esun_constituent_snapshot(etf_code: str, *, timeout_seconds: float = 30,
                                    fetched_at: datetime | None = None):
    code = _normalize_code(etf_code)
    overview = _post_json(f"{ESUN_API_BASE_URL}/GetETFOverview", {
        "PageSize": 999, "ETFInvestmentIDs": [], "ETFFundTypeIDs": [],
        "ETFFundDividendIDs": [], "ETFInvestAreaIDs": [], "Keyword": code,
        "KeywordId": 0, "SortColName": "", "IsDesc": False,
    }, timeout_seconds=timeout_seconds)
    fund_id = parse_esun_mapping(etf_code=code, overview_payload=overview)
    assets = _post_json(
        f"{ESUN_API_BASE_URL}/GetFundAssets",
        {"FundID": fund_id, "SearchDate": None}, timeout_seconds=timeout_seconds,
    )
    return parse_esun_constituent_payload(
        assets, etf_code=code, fund_id=fund_id, source_url=ESUN_SOURCE_URL,
        fetched_at=fetched_at or datetime.now(timezone.utc),
    )


def fetch_franklin_constituent_snapshot(
    etf_code: str, *, timeout_seconds: float = 30,
    fetched_at: datetime | None = None,
) -> ETFConstituentSnapshotCreate:
    code = _normalize_code(etf_code)
    catalog = _get(f"{FRANKLIN_API_BASE_URL}/etf", timeout_seconds=timeout_seconds)
    fund_id = parse_franklin_mapping(
        etf_code=code, catalog_payload=catalog.json()
    )
    dates = _get(
        f"{FRANKLIN_API_BASE_URL}/etf/share-dates/{fund_id}",
        timeout_seconds=timeout_seconds,
    )
    query_date = parse_franklin_latest_query_date(dates.json())
    holdings = _get(
        f"{FRANKLIN_API_BASE_URL}/etf/shares/{fund_id}?date={query_date}",
        timeout_seconds=timeout_seconds,
    )
    return parse_franklin_constituent_payload(
        holdings.json(), etf_code=code, fund_id=fund_id,
        source_url=f"{FRANKLIN_SOURCE_URL}?id={fund_id}&tab=holdings",
        fetched_at=fetched_at or datetime.now(timezone.utc),
    )
MAPPED_CONSTITUENT_FETCHERS: dict[str, Callable[..., ETFConstituentSnapshotCreate]] = {
    "mega": fetch_mega_constituent_snapshot,
    "fuh_hwa": fetch_fuh_hwa_constituent_snapshot,
    "uob": fetch_uob_constituent_snapshot,
    "esun": fetch_esun_constituent_snapshot,
    "franklin": fetch_franklin_constituent_snapshot,
}
