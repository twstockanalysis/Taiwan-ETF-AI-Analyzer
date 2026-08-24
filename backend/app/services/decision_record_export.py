"""M11-4 決策紀錄 Excel 匯出。"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.models.decision_profile import DecisionRecordResponse


_DARK_BLUE = "17365D"
_MID_BLUE = "D9EAF7"
_LIGHT_BLUE = "EDF4FA"
_WHITE = "FFFFFF"
_THIN_GRAY = Side(style="thin", color="B7C9D6")
_CURRENCY_FORMAT = '#,##0.00 "TWD"'
_PERCENT_FORMAT = '0.00"%"'


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime, int, float)) or value is None:
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _style_title(sheet, title: str, subtitle: str) -> None:
    sheet.merge_cells("A1:D1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color=_WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=_DARK_BLUE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells("A2:D2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color="44546A")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[2].height = 34


def _write_key_values(
    sheet,
    start_row: int,
    title: str,
    rows: Iterable[tuple[str, Any, str | None]],
) -> int:
    sheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=4,
    )
    cell = sheet.cell(start_row, 1, title)
    cell.font = Font(bold=True, color=_DARK_BLUE)
    cell.fill = PatternFill("solid", fgColor=_MID_BLUE)
    row = start_row + 1
    for label, value, number_format in rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 1).font = Font(bold=True, color="44546A")
        value_cell = sheet.cell(row, 2, _excel_value(value))
        if number_format:
            value_cell.number_format = number_format
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        for column in range(1, 5):
            target = sheet.cell(row, column)
            target.border = Border(bottom=_THIN_GRAY)
            target.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1


def _write_table(
    sheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, values in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, _excel_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=_THIN_GRAY)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    return start_row + len(rows) + 2


def _finalize_sheet(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 48)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and cell.alignment == Alignment():
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _comparison_sheet(workbook: Workbook, record: DecisionRecordResponse) -> None:
    sheet = workbook.create_sheet("分析比較")
    _style_title(
        sheet,
        "候選加入前後分析比較",
        "所有數字均來自保存當下的分析快照；空白代表無法計算，不代表零。",
    )
    comparison = record.analysis.comparison
    fields = [
        ("投入後總值", "total_value_before", "total_value_after", _CURRENCY_FORMAT),
        (
            "年均稅後可用現金",
            "annual_after_tax_cash_before",
            "annual_after_tax_cash_after",
            _CURRENCY_FORMAT,
        ),
        (
            "年度目標覆蓋率",
            "target_coverage_pct_before",
            "target_coverage_pct_after",
            _PERCENT_FORMAT,
        ),
        (
            "資金缺口",
            "funding_shortfall_before",
            "funding_shortfall_after",
            _CURRENCY_FORMAT,
        ),
        (
            "稅後總報酬率",
            "after_tax_total_return_pct_before",
            "after_tax_total_return_pct_after",
            _PERCENT_FORMAT,
        ),
    ]
    rows = []
    for label, before, after, number_format in fields:
        before_value = getattr(comparison, before) if comparison else None
        after_value = getattr(comparison, after) if comparison else None
        delta = (
            after_value - before_value
            if before_value is not None and after_value is not None
            else None
        )
        rows.append([label, before_value, after_value, delta])
    _write_table(sheet, 4, ["比較項目", "目前持倉", "加入候選後", "變化"], rows)
    for row in range(5, 5 + len(rows)):
        number_format = fields[row - 5][3]
        for column in range(2, 5):
            sheet.cell(row, column).number_format = number_format
    _finalize_sheet(sheet, [26, 20, 20, 20])


def _notes_sheet(workbook: Workbook, record: DecisionRecordResponse) -> None:
    sheet = workbook.create_sheet("理由與風險")
    _style_title(
        sheet,
        "理由、排除、替代方案與風險註記",
        "穩定代碼與原始訊息一併保存，便於日後追溯當時判定。",
    )
    rows = []
    for category, notes in (
        ("採用理由", record.rationale),
        ("排除理由", record.exclusions),
        ("可行替代方案", record.alternatives),
        ("風險註記", record.risk_notes),
    ):
        rows.extend(
            [
                category,
                note.code,
                note.message,
                ", ".join(map(str, note.affected_months)),
            ]
            for note in notes
        )
    _write_table(sheet, 4, ["類別", "代碼", "說明", "影響月份"], rows)
    _finalize_sheet(sheet, [18, 34, 48, 16])


def _holdings_sheet(workbook: Workbook, record: DecisionRecordResponse) -> None:
    sheet = workbook.create_sheet("持倉快照")
    _style_title(
        sheet,
        "目前與候選加入後持倉快照",
        "價格與單位均為保存當下情境；參考價格不是即時報價。",
    )
    rows = []
    for snapshot_label, portfolio in (
        ("目前持倉", record.analysis.current_portfolio),
        ("加入候選後", record.analysis.proposed_portfolio),
    ):
        if portfolio is None:
            continue
        for holding in portfolio.holdings:
            rows.append(
                [
                    snapshot_label,
                    holding.etf_code,
                    holding.name,
                    holding.held_units,
                    holding.unit_price,
                    holding.current_value,
                    holding.annual_gross_distribution_cash,
                    holding.price_return_period_code,
                    holding.annualized_price_return_pct,
                ]
            )
    _write_table(
        sheet,
        4,
        [
            "快照",
            "ETF 代號",
            "名稱",
            "單位數",
            "參考單價",
            "部位價值",
            "年均稅前配息現金",
            "報酬期間",
            "年化價格報酬率",
        ],
        rows,
    )
    for row in range(5, 5 + len(rows)):
        sheet.cell(row, 4).number_format = "#,##0"
        for column in (5, 6, 7):
            sheet.cell(row, column).number_format = _CURRENCY_FORMAT
        sheet.cell(row, 9).number_format = _PERCENT_FORMAT
    _finalize_sheet(sheet, [16, 14, 24, 14, 18, 18, 22, 16, 20])


def _inputs_sheet(workbook: Workbook, record: DecisionRecordResponse) -> None:
    sheet = workbook.create_sheet("限制與輸入")
    _style_title(
        sheet,
        "輸入、限制與資料缺口",
        "此頁保留使用者輸入與無法取得欄位，避免把未知值誤當成零。",
    )
    request = record.request
    eligibility = record.analysis.eligibility
    evaluated_candidates = (
        [
            *eligibility.selected_candidates,
            *eligibility.rejected_candidates,
        ]
        if eligibility is not None
        else []
    )
    automatic_overlap = (
        evaluated_candidates[0].holding_overlap_pct
        if evaluated_candidates
        else None
    )
    row = _write_key_values(
        sheet,
        4,
        "候選輸入",
        [
            ("候選 ETF", record.candidate_etf_code, None),
            ("預計增加單位數", request.proposed_units, "#,##0"),
            ("候選參考單價", request.unit_price, _CURRENCY_FORMAT),
            ("自動成分股重疊", automatic_overlap, _PERCENT_FORMAT),
            ("月配缺口判定", "啟用" if request.monthly_coverage_enabled else "停用", None),
        ],
    )
    unavailable = [
        [item.field, item.reason]
        for item in record.analysis.unavailable_fields
    ] or [["無", "本次快照沒有頂層資料缺口。"]]
    row = _write_table(sheet, row, ["無法取得欄位", "原因"], unavailable)
    priority_rows = [
        [index, priority]
        for index, priority in enumerate(record.analysis.decision_priority, start=1)
    ]
    _write_table(sheet, row, ["判定順序", "固定優先項目"], priority_rows)
    _finalize_sheet(sheet, [28, 48, 18, 18])


def export_decision_record_xlsx(record: DecisionRecordResponse) -> bytes:
    """將單一不可變紀錄輸出為可稽核的 Excel 活頁簿。"""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "決策摘要"
    _style_title(
        summary,
        "ETF 候選分析決策紀錄",
        "不可變分析快照；不是投資建議、報酬保證或交易指示。",
    )
    comparison = record.analysis.comparison
    _write_key_values(
        summary,
        4,
        "紀錄資訊",
        [
            ("紀錄編號", record.id, "0"),
            ("建立時間", record.created_at, "yyyy-mm-dd hh:mm:ss"),
            ("候選 ETF", f"{record.candidate_etf_code} {record.candidate_name}", None),
            ("分析狀態", record.analysis_status.value, None),
            ("資格結果", record.outcome, None),
            ("分析日期", record.analysis.analysis_date, "yyyy-mm-dd"),
            (
                "候選投入金額",
                comparison.additional_capital if comparison else None,
                _CURRENCY_FORMAT,
            ),
            ("估算標示", record.analysis.estimate_label, None),
            ("紀錄特性", "不可變；後續重算會新增紀錄", None),
            ("券商與交易", "未連接券商；不送出交易", None),
        ],
    )
    _finalize_sheet(summary, [24, 28, 22, 22])
    _comparison_sheet(workbook, record)
    _notes_sheet(workbook, record)
    _holdings_sheet(workbook, record)
    _inputs_sheet(workbook, record)
    workbook.properties.title = f"ETF decision record {record.id}"
    workbook.properties.subject = "Taiwan ETF candidate scenario snapshot"
    workbook.properties.creator = "ETF奈米戶"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
