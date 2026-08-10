"""M11-4 不可變決策紀錄與 Excel 匯出整合測試。"""

from io import BytesIO
from pathlib import Path
import tempfile
import unittest

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.api.dependencies import get_database_path
from backend.app.database.connection import get_connection
from backend.app.database.init_db import initialize_database
from backend.app.main import create_app


class TestDecisionRecord(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.database_path = Path(self.temp_directory.name) / "decision-record.db"
        initialize_database(self.database_path)
        connection = get_connection(self.database_path)
        connection.executemany(
            """
            INSERT INTO etf_master (code, name, is_active, is_bond)
            VALUES (?, ?, 0, 0);
            """,
            [("0056", "元大高股息"), ("00878", "國泰永續高股息")],
        )
        connection.commit()
        connection.close()
        self.application = create_app()
        self.application.dependency_overrides[get_database_path] = (
            lambda: self.database_path
        )
        self.client = TestClient(self.application)
        self.client.put(
            "/api/v1/decision-profile/conditions",
            json={
                "monthly_after_tax_target": 3000,
                "analysis_years": 10,
                "history_years": 3,
                "cash_deduction_rate_pct": 10,
            },
        )
        self.client.put(
            "/api/v1/decision-profile/holdings/0056",
            json={"held_units": 1000, "unit_price": 30},
        )

    def tearDown(self):
        self.client.close()
        self.application.dependency_overrides.clear()
        self.temp_directory.cleanup()

    @staticmethod
    def payload():
        return {
            "proposed_units": 100,
            "unit_price": 20,
            "holding_overlap_pct": None,
            "monthly_coverage_enabled": True,
        }

    def _create_record(self):
        return self.client.post(
            "/api/v1/decision-profile/candidate-analysis/00878/decision-records",
            json=self.payload(),
        )

    def test_openapi_exposes_create_read_list_and_export_only(self):
        paths = self.application.openapi()["paths"]
        collection = "/api/v1/decision-profile/decision-records"
        detail = "/api/v1/decision-profile/decision-records/{record_id}"
        export = detail + "/export.xlsx"
        create = (
            "/api/v1/decision-profile/candidate-analysis/{etf_code}"
            "/decision-records"
        )
        self.assertEqual(set(paths[collection]), {"get"})
        self.assertEqual(set(paths[detail]), {"get"})
        self.assertEqual(set(paths[export]), {"get"})
        self.assertEqual(set(paths[create]), {"post"})

    def test_create_preserves_reasons_alternatives_and_risks(self):
        response = self._create_record()

        self.assertEqual(response.status_code, 201)
        record = response.json()
        self.assertTrue(record["immutable"])
        self.assertFalse(record["broker_connected"])
        self.assertEqual(record["candidate_etf_code"], "00878")
        self.assertEqual(record["request"]["proposed_units"], 100)
        self.assertEqual(record["analysis"]["candidate_etf_code"], "00878")
        self.assertTrue(record["exclusions"])
        self.assertTrue(record["alternatives"])
        risk_codes = {item["code"] for item in record["risk_notes"]}
        self.assertIn("USER_ENTERED_REFERENCE_PRICE", risk_codes)
        self.assertIn("IMMUTABLE_SNAPSHOT", risk_codes)

        listed = self.client.get(
            "/api/v1/decision-profile/decision-records"
        ).json()
        self.assertEqual([item["id"] for item in listed], [record["id"]])

    def test_saved_snapshot_does_not_change_with_profile(self):
        record = self._create_record().json()
        self.client.put(
            "/api/v1/decision-profile/conditions",
            json={
                "monthly_after_tax_target": 9999,
                "analysis_years": 5,
                "history_years": 2,
                "cash_deduction_rate_pct": None,
            },
        )

        saved = self.client.get(
            f"/api/v1/decision-profile/decision-records/{record['id']}"
        ).json()

        self.assertEqual(
            saved["analysis"]["current_portfolio"]["conditions"]
            ["monthly_after_tax_target"],
            "3000.0",
        )
        self.assertEqual(saved["request"]["unit_price"], "20")

    def test_export_is_readable_workbook_with_typed_values(self):
        record = self._create_record().json()
        response = self.client.get(
            f"/api/v1/decision-profile/decision-records/{record['id']}"
            "/export.xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["決策摘要", "分析比較", "理由與風險", "持倉快照", "限制與輸入"],
        )
        summary = workbook["決策摘要"]
        self.assertEqual(summary["A1"].value, "ETF 候選分析決策紀錄")
        self.assertEqual(summary["B5"].value, record["id"])
        self.assertIsInstance(summary["B11"].value, (int, float))
        notes = workbook["理由與風險"]
        note_codes = {
            notes.cell(row, 2).value
            for row in range(5, notes.max_row + 1)
        }
        self.assertIn("IMMUTABLE_SNAPSHOT", note_codes)
        self.assertNotIn("Sheet", workbook.sheetnames)

    def test_unknown_record_and_candidate_return_404(self):
        self.assertEqual(
            self.client.get(
                "/api/v1/decision-profile/decision-records/999"
            ).status_code,
            404,
        )
        response = self.client.post(
            "/api/v1/decision-profile/candidate-analysis/9999/decision-records",
            json=self.payload(),
        )
        self.assertEqual(response.status_code, 404)


if __name__ == "__main__":
    unittest.main()
